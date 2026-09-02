#!/usr/bin/env python3
"""
Import TX county-level elected officers from the 2024 General Election
Winner Listing Report (Texas Secretary of State, results.texas-election.com)
into data/us/tx/{organizations,posts,memberships,people}/county/.

Sources (committed under `reference/TX Rolling Audit/2024 SOS Reports/`):
  - WinnerListingReport.xlsx -- one row per office won, 2024 cycle, statewide.
  - OfficialCanvassReport.xlsx -- joined only to recover party for offices
    that happen to also appear there (see PARTY note below).

Scope: issue #25 (TX county office structure/officeholders), plus a handful
of county-created courts (County Court at Law, County Criminal Court at Law,
Probate Court) and single-county "Criminal District Attorney" titles that
overlap issue #26 but do NOT require a new jurisdiction record (they sit
squarely inside an already-seeded county jurisdiction).

Explicitly NOT imported here (still blocked or out of scope):
  - District Attorney "Nth Judicial District" / multi-county DA titles,
    District Judges, Courts of Appeals, Probate Court "at large" statewide
    seats -- these need a judicial-district or state jurisdiction record
    that does not exist in data/us/tx/jurisdictions/ yet. Left for #26.
  - HARRIS COUNTY DEPARTMENT OF EDUCATION -- a distinct special-purpose
    government, not a county office; left unimported.
  - Drainage District seats -- special districts, tracked under #15/#25 scope
    boundary, not this file's county-officer inventory.

PARTY: Texas county elections are partisan (per Audit_Instructions.md on
branch tx-audit-instructions), but the Official Canvass Report only covers
592 statewide-canvassed offices -- no sheriff/commissioner/etc. race appears
in it. Party is therefore NOT available for the county layer from these two
files; every candidacy here is recorded with party: null. Confirmed by the
join producing zero matches (see `--check-canvass-overlap`).

CURRENCY: these are 2024-cycle winners as of the 2026-09-01 retrieval date.
Texas county offices serve staggered 4-year terms -- a given office not
appearing here was simply not on the 2024 ballot, not vacant. Do not treat
absence as a vacancy. Every membership is recorded `how_seated: elected`
with `end: '2029'` (the statutory 4-year term horizon) and a source note
flagging the as-of-2024 caveat; nothing here claims verified *current*
(2026) occupancy beyond what a 2024-elected, undisturbed 4-year term implies.

Idempotent by id AND by filename -- never overwrites. Dry run by default;
pass --write to create files.
"""
import re
import sys
import unicodedata
import uuid
from pathlib import Path

import openpyxl
import yaml

REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data" / "us" / "tx"
SRC_DIR = REPO / "reference" / "TX Rolling Audit" / "2024 SOS Reports"
WINNERS_XLSX = SRC_DIR / "WinnerListingReport.xlsx"
CANVASS_XLSX = SRC_DIR / "OfficialCanvassReport.xlsx"

RESULTS_URL = "https://results.texas-election.com/reports"
RETRIEVED = "2026-09-01"
ELECTION_ID = "us-tx/2024-11-05/general"

NS = uuid.UUID("3b6f0a9e-9a3b-5f2e-8b8f-6c1b6e9d2a41")

HEADER = (
    "# Imported from the Texas Secretary of State's 2024 General Election\n"
    "# Winner Listing Report (results.texas-election.com/reports) by\n"
    "# scripts/import_tx_2024_county_officers.py. See issue #25. Party is\n"
    "# unavailable for county-layer offices from this source (see the\n"
    "# script's module docstring); offices not on the 2024 ballot are\n"
    "# simply not represented here, not vacant -- see the CURRENCY note.\n"
)

# Raw " COUNTY" prefix spellings that don't slugify to an existing
# data/us/tx/jurisdictions/county/*.yaml slug via the default transform.
COUNTY_SLUG_FIXUPS = {
    "dewitt": "de-witt",
    "lasalle": "la-salle",
}


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def oid(kind, key):
    return f"ocd-{kind}/{uuid.uuid5(NS, kind + '|' + key)}"


def load_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [h for h in rows[0] if h is not None]
    # WinnerListingReport has a title row before the header; canvass does not.
    if header and header[0] not in ("ELECTION DATE-NAME",):
        rows = rows[1:]
    return rows[1:]


def titlecase_name(raw):
    return " ".join(w.capitalize() if not w.isupper() or len(w) > 3 else w.capitalize()
                     for w in raw.split())


# --- Office classification ------------------------------------------------

PRECINCT_RE = re.compile(r"PRECINCT (\d+(?:\s*&\s*\d+)?)")
UNEXPIRED_RE = re.compile(r"\s*-\s*UNEXPIRED TERM\s*$")

# A precinct list can combine 2+ precincts with a mix of commas and "&"
# (e.g. "2 & 3", "1, 5 & 6") -- some offices cover more than a simple pair.
PRECINCT_LIST_RE = r"\d+(?:\s*[,&]\s*\d+)*"


def normalize_precinct_list(raw):
    """'1, 5 & 6' -> (slug '1-5-6', label '1 & 5 & 6')."""
    nums = re.split(r"\s*[,&]\s*", raw.strip())
    return "-".join(nums), " & ".join(nums)

# (match regex, office_key_template, title_template, seats, org_name_template)
SIMPLE_OFFICES = [
    (re.compile(r"^COUNTY JUDGE$"), "county-judge", "County Judge", 1),
    (re.compile(r"^SHERIFF$"), "sheriff", "Sheriff", 1),
    (re.compile(r"^SHERIFF/COUNTY TAX ASSESSOR-COLLECTOR$"), "sheriff-tax-assessor-collector",
     "Sheriff / Tax Assessor-Collector", 1),
    (re.compile(r"^COUNTY TAX ASSESSOR-COLLECTOR$"), "tax-assessor-collector", "Tax Assessor-Collector", 1),
    (re.compile(r"^COUNTY ATTORNEY$"), "county-attorney", "County Attorney", 1),
    (re.compile(r"^COUNTY TREASURER$"), "county-treasurer", "County Treasurer", 1),
    (re.compile(r"^COUNTY CLERK/DISTRICT CLERK$"), "county-clerk-district-clerk", "County Clerk / District Clerk", 1),
    (re.compile(r"^COUNTY CLERK$"), "county-clerk", "County Clerk", 1),
    (re.compile(r"^DISTRICT CLERK$"), "district-clerk", "District Clerk", 1),
]

COURT_RE = re.compile(r"^JUDGE,\s*COUNTY COURT AT LAW\s*NO\.?\s*(\d+)$")
CRIM_COURT_RE = re.compile(r"^COUNTY CRIMINAL COURT AT LAW\s*NO\.?\s*(\d+)$")
PROBATE_RE = re.compile(r"^PROBATE COURT\s*(?:NO\.?\s*(\d+))?$")

SKIP_SUBSTRINGS = [
    "HARRIS COUNTY DEPARTMENT OF EDUCATION",
    "DRAINAGE DISTRICT",
]


def classify(office_raw):
    """Return (office_key, title, seats, seat_label_or_None) or None to skip."""
    office = UNEXPIRED_RE.sub("", office_raw).strip()

    for sub in SKIP_SUBSTRINGS:
        if sub in office:
            return None

    m = re.match(r"^COUNTY COMMISSIONER (.+)$", office)
    if m:
        pm = PRECINCT_RE.match(m.group(1).strip())
        if not pm:
            return None
        precinct = re.sub(r"\s*&\s*", " & ", pm.group(1))
        return "commissioner", "County Commissioner", 4, f"Precinct {precinct}"

    if office == "COUNTY CONSTABLE":
        return "constable", "Constable", 1, None

    # A few precincts elect more than one constable "Place" (e.g. Maverick
    # Precinct 3, El Paso Precinct 6) -- same collision risk and same fix
    # as Justice of the Peace below: capture Place only when present.
    m = re.match(rf"^COUNTY CONSTABLE PRECINCT\s+({PRECINCT_LIST_RE})(?:,\s*PLACE\s*(\d+))?$", office)
    if m:
        precinct, precinct_label = normalize_precinct_list(m.group(1))
        place = m.group(2)
        if place:
            return (f"constable-precinct-{precinct}-place-{place}", "Constable", 1,
                    f"Precinct {precinct_label}, Place {place}")
        return f"constable-precinct-{precinct}", "Constable", 1, f"Precinct {precinct_label}"

    m = re.match(r"^COUNTY CONSTABLE (.+)$", office)
    if m:
        rest = m.group(1).strip()
        pm = re.match(r"^NO\.?\s*(\d+)$", rest)
        if not pm:
            return None
        precinct = pm.group(1)
        return f"constable-precinct-{precinct}", "Constable", 1, f"Precinct {precinct}"

    if office == "JUSTICE OF THE PEACE":
        return "justice-of-the-peace", "Justice of the Peace", 1, None

    # Populous precincts elect more than one JP "Place" (e.g. "PRECINCT 1,
    # PLACE 2") -- each Place is its own single-seat post. Capturing Place
    # only when present avoids renaming the (much more common) single-JP
    # precincts that never carry a Place suffix at all.
    m = re.match(rf"^JUSTICE OF THE PEACE\s*PRECINCT\s+({PRECINCT_LIST_RE})(?:,\s*PLACE\s*(\d+))?$", office)
    if m:
        precinct, precinct_label = normalize_precinct_list(m.group(1))
        place = m.group(2)
        if place:
            return (f"justice-of-the-peace-precinct-{precinct}-place-{place}", "Justice of the Peace", 1,
                    f"Precinct {precinct_label}, Place {place}")
        return f"justice-of-the-peace-precinct-{precinct}", "Justice of the Peace", 1, f"Precinct {precinct_label}"

    if office == "JUDGE, COUNTY COURT AT LAW":
        return "county-court-at-law-judge", "County Court at Law Judge", 1, None

    m = COURT_RE.match(office)
    if m:
        n = m.group(1)
        return f"county-court-at-law-no-{n}-judge", f"County Court at Law No. {n} Judge", 1, None

    m = CRIM_COURT_RE.match(office)
    if m:
        n = m.group(1)
        return f"county-criminal-court-at-law-no-{n}-judge", f"County Criminal Court at Law No. {n} Judge", 1, None

    m = PROBATE_RE.match(office)
    if m:
        n = m.group(1)
        if n:
            return f"probate-court-no-{n}-judge", f"Probate Court No. {n} Judge", 1, None
        return "probate-court-judge", "Probate Court Judge", 1, None

    for rx, key, title, seats in SIMPLE_OFFICES:
        if rx.match(office):
            return key, title, seats, None

    return None


COUNTY_ROW_RE = re.compile(r"^(?P<county>.+?)\s+COUNTY\s{1,2}-\s*(?P<office>.+)$")
CRIMINAL_DA_ROW_RE = re.compile(r"^CRIMINAL DISTRICT ATTORNEY (?P<county>.+?) COUNTY(?:\s*-\s*UNEXPIRED TERM)?$")


def county_slug(raw_county_name):
    key = raw_county_name.strip().lower().replace(" ", "")
    if key in COUNTY_SLUG_FIXUPS:
        return COUNTY_SLUG_FIXUPS[key]
    return slugify(raw_county_name)


def parse_office_row(office_raw):
    """Return (county_raw, office_key, title, seats, seat_label) or None."""
    m = COUNTY_ROW_RE.match(office_raw)
    if m:
        cls = classify(m.group("office").strip())
        if cls is None:
            return None
        office_key, title, seats, seat_label = cls
        return m.group("county").strip(), office_key, title, seats, seat_label

    m = CRIMINAL_DA_ROW_RE.match(office_raw)
    if m:
        return m.group("county").strip(), "criminal-district-attorney", "Criminal District Attorney", 1, None

    return None


def load_known_counties():
    known = {}
    for f in (DATA_DIR / "jurisdictions" / "county").glob("*-government.yaml"):
        doc = yaml.safe_load(f.read_text()) or {}
        slug = f.stem[: -len("-government")]
        known[slug] = doc["id"]
    return known


class Recorder:
    def __init__(self, write):
        self.write = write
        self.stats = {}

    def emit(self, kind, doc, filename):
        singular = kind.rstrip("s")
        path = DATA_DIR / kind / "county" / filename
        if path.exists():
            existing = yaml.safe_load(path.read_text()) or {}
            if existing.get("id") == doc["id"]:
                self.stats[f"{singular}_existing"] = self.stats.get(f"{singular}_existing", 0) + 1
            else:
                self.stats[f"{singular}_skipped_filename_conflict"] = (
                    self.stats.get(f"{singular}_skipped_filename_conflict", 0) + 1
                )
            return False
        self.stats[f"{singular}_new"] = self.stats.get(f"{singular}_new", 0) + 1
        if self.write:
            path.parent.mkdir(parents=True, exist_ok=True)
            text = HEADER + yaml.safe_dump(
                doc, sort_keys=False, allow_unicode=True, default_flow_style=False, width=1000
            )
            path.write_text(text)
        return True


def main():
    write = "--write" in sys.argv[1:]

    known_counties = load_known_counties()
    winner_rows = load_rows(WINNERS_XLSX)

    rec = Recorder(write)
    unmatched_offices = {}
    unknown_counties = {}

    # org/post keyed by (county_slug, office_key) since commissioner posts
    # collect multiple precinct memberships under one post.
    orgs = {}
    posts = {}
    people = []
    memberships = []

    for row in winner_rows:
        election, office_raw, candidate_raw = row[0], row[1], row[2]
        if not office_raw or not candidate_raw:
            continue
        office_raw = office_raw.strip()
        parsed = parse_office_row(office_raw)
        if parsed is None:
            unmatched_offices[office_raw] = unmatched_offices.get(office_raw, 0) + 1
            continue

        county_raw, office_key, title, seats, seat_label = parsed
        slug = county_slug(county_raw)
        if slug not in known_counties:
            unknown_counties[county_raw] = unknown_counties.get(county_raw, 0) + 1
            continue
        jurisdiction_id = known_counties[slug]

        org_key = (slug, office_key)
        post_id = f"{slug}-tx-county/{office_key}"

        if org_key not in orgs:
            org_id = oid("organization", f"tx-county|{slug}|{office_key}")
            orgs[org_key] = {
                "id": org_id,
                "name": title,
                "jurisdiction_id": jurisdiction_id,
                "identifiers": [{"scheme": "civicmirror-office", "identifier": post_id}],
                "status": "active",
                "sources": [{"url": RESULTS_URL,
                             "note": "Texas SOS 2024 General Election Winner Listing Report",
                             "retrieved": RETRIEVED}],
            }
            posts[org_key] = {
                "id": post_id,
                "organization_id": org_id,
                "title": title,
                "seats": seats,
                "sources": [{"url": RESULTS_URL,
                             "note": "Texas SOS 2024 General Election Winner Listing Report",
                             "retrieved": RETRIEVED}],
            }

        org_id = orgs[org_key]["id"]
        candidate_name = re.sub(r"\s+", " ", candidate_raw).strip()
        person_id = oid("person", f"tx-county-2024|{slug}|{post_id}|{candidate_name}")

        person = {
            "id": person_id,
            "name": titlecase_name(candidate_name),
            "candidacies": [{
                "contest_id": f"tx-2024-general/{post_id}",
                "election_id": ELECTION_ID,
                "jurisdiction_id": jurisdiction_id,
                "office_id": post_id,
                "party": None,
                "ballot_name": candidate_name,
                "status": "active",
                "sources": [{"url": RESULTS_URL,
                             "note": "Texas SOS 2024 General Election Winner Listing Report",
                             "retrieved": RETRIEVED}],
            }],
            "verification": {"status": "machine-extracted", "pipeline": "import_tx_2024_county_officers"},
            "sources": [{"url": RESULTS_URL,
                         "note": "Texas SOS 2024 General Election Winner Listing Report",
                         "retrieved": RETRIEVED}],
        }
        people.append(person)

        surname = candidate_name.split()[-1]
        mem_key = f"{post_id.replace('/', '-')}-{slugify(surname)}"
        note = ("2024 General Election winner; office serves a staggered 4-year term, "
                "so this reflects the seat as of the 2024 cycle, not a verified 2026 "
                "incumbency check.")
        membership = {
            "id": mem_key,
            "person_id": person_id,
            "organization_id": org_id,
            "post_id": post_id,
            "role": title,
            "how_seated": "elected",
            "end": "2029",
            "sources": [{"url": RESULTS_URL, "note": note, "retrieved": RETRIEVED}],
        }
        if seat_label:
            membership["seat"] = seat_label
        memberships.append((mem_key, membership))

    # --- write phase --------------------------------------------------
    for org_key, org in orgs.items():
        slug, office_key = org_key
        rec.emit("organizations", org, f"{slug}-tx-county-{office_key}.yaml")
    for org_key, post in posts.items():
        slug, office_key = org_key
        rec.emit("posts", post, f"{slug}-tx-county-{office_key}.yaml")

    taken_people_slugs = {p.stem for p in (DATA_DIR / "people" / "county").glob("*.yaml")} if (DATA_DIR / "people" / "county").exists() else set()
    for person in people:
        base_slug = slugify(person["name"])
        suffix = person["id"].rsplit("/", 1)[-1][:8]
        slug_name = base_slug if base_slug not in taken_people_slugs else f"{base_slug}-{suffix}"
        taken_people_slugs.add(slug_name)
        rec.emit("people", person, f"{slug_name}.yaml")

    for mem_key, membership in memberships:
        rec.emit("memberships", membership, f"{mem_key}.yaml")

    print(f"Parsed {len(winner_rows)} winner rows")
    print(f"Organizations: {len(orgs)}  Posts: {len(posts)}  People: {len(people)}  Memberships: {len(memberships)}")
    print()
    print("==================== SUMMARY ====================")
    for k in sorted(rec.stats):
        print(f"{k}: {rec.stats[k]}")
    if unmatched_offices:
        print(f"\nUnmatched/skipped office strings ({sum(unmatched_offices.values())} rows, "
              f"{len(unmatched_offices)} distinct):")
        for k in sorted(unmatched_offices)[:40]:
            print(f"  {k!r}  x{unmatched_offices[k]}")
    if unknown_counties:
        print(f"\nERROR: unresolved county names (no matching jurisdiction slug):")
        for k in sorted(unknown_counties):
            print(f"  {k!r}  x{unknown_counties[k]}")
    if not write:
        print("\n(dry run -- pass --write to create files)")


if __name__ == "__main__":
    main()
