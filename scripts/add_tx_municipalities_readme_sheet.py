#!/usr/bin/env python3
"""
Add a README sheet to the front of TX_Municipalities.xlsx documenting what
each sheet tracks and its current coverage, computed live from the
workbook rather than hardcoded, so it can be re-run after future updates.

Inserts (or replaces) the sheet at index 0. Leaves every other sheet's
cells and formatting untouched.

Usage: python3 add_tx_municipalities_readme_sheet.py [--write]
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

REPO = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO / "reference" / "TX Rolling Audit" / "TX_Municipalities.xlsx"


def count_data_rows(ws, header_row):
    n = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            n += 1
    return n


def col_value(ws, header_row, col_idx, data_row_offset=0):
    return [r[col_idx] for r in ws.iter_rows(min_row=header_row + 1 + data_row_offset, values_only=True)]


def build_rows(wb):
    """Returns list of (sheet, header_row_index, description, scope_lines)."""

    counties = wb["Counties"]
    n_counties = count_data_rows(counties, 1)
    n_counties_site = sum(1 for r in counties.iter_rows(min_row=2, values_only=True) if r[0] and r[1])

    muni = wb["Municipalities"]
    n_muni = count_data_rows(muni, 1)
    n_muni_site = sum(1 for r in muni.iter_rows(min_row=2, values_only=True) if r[0] and (r[6] or r[11]))

    isd = wb["ISD"]
    n_isd = count_data_rows(isd, 1)
    n_isd_elected = sum(1 for r in isd.iter_rows(min_row=2, values_only=True) if r[0] and r[10] == "Elected")
    n_isd_appointed = sum(1 for r in isd.iter_rows(min_row=2, values_only=True) if r[0] and r[10] == "Appointed (excluded)")
    n_isd_gap = sum(1 for r in isd.iter_rows(min_row=2, values_only=True) if r[0] and r[9] and "needs sourcing" in r[9])

    ccd = wb["CCD"]
    n_ccd = count_data_rows(ccd, 1)

    spd = wb["Special Districts (Non-MUD)"]
    n_spd = count_data_rows(spd, 1)
    n_spd_active = sum(1 for r in spd.iter_rows(min_row=2, values_only=True) if r[0] and r[3] == "ACTIVE")

    mud = wb["MUD"]
    n_mud = count_data_rows(mud, 1)
    n_mud_active = sum(1 for r in mud.iter_rows(min_row=2, values_only=True) if r[0] and r[2] == "ACTIVE")

    jud = wb["Judiciary"]
    n_jud = count_data_rows(jud, 2)
    from collections import Counter
    jud_types = Counter(r[1] for r in jud.iter_rows(min_row=3, values_only=True) if r[0])
    jud_summary = "; ".join(f"{k}: {v}" for k, v in sorted(jud_types.items()))

    swcd = wb["Soil & Water Conservation"]
    n_swcd = count_data_rows(swcd, 2)

    apd = wb["Appraisal Districts"]
    n_apd = count_data_rows(apd, 1)
    n_apd_elected = sum(1 for r in apd.iter_rows(min_row=2, values_only=True) if r[0] and r[5] and r[5].startswith("Yes"))

    tml = wb["TML Title Codes"]
    n_tml = count_data_rows(tml, 1)

    return [
        ("TML Title Codes",
         "Reference lookup, not an enumeration: every office-title code used in the Texas Municipal League "
         "member directory (directory.tml.org), with a flag for whether that title is typically an elected "
         "office. Used to scope which TML-listed positions belong in an elected-office audit.",
         f"{n_tml} title codes."),

        ("Counties",
         "Starting enumeration for issue #25 (TX county office structure/officeholders): all 254 Texas counties "
         "seeded from the U.S. Census Bureau's 2022 Census of Governments, with each county's official website. "
         "Office-level structure and officeholder data lives in the repo's data files "
         "(data/us/tx/{organizations,posts,memberships,people}/county/), not in this sheet.",
         f"{n_counties} counties, {n_counties_site} with a website on file (100%)."),

        ("Municipalities",
         "Classification and source-of-record tracking for issue #10: every Texas municipality with its county, "
         "TML region, population, general-law type or home-rule status, and website/municipal-code links. "
         "The starting point for the classify-then-template municipal audit described in Audit_Instructions.md.",
         f"{n_muni} municipalities, {n_muni_site} with a verified website ({n_muni_site*100//n_muni}%)."),

        ("ISD",
         "Independent school district structure and officeholders for issue #13, seeded from the Texas "
         "Education Agency's AskTED district directory and BBB(LOCAL) election-method policies via TASB. "
         "Jurisdiction/Board Size/Seat Structure/Officeholders Seeded/Status columns cross-reference "
         "data/us/tx/{jurisdictions,organizations,posts,people,memberships}/school/ -- computed live from "
         "those files, not hand-maintained. 5 districts (Randolph Field, Lackland, Ft Sam Houston, Boys Ranch, "
         "and Houston ISD under its current TEA receivership) are correctly excluded as appointed, not elected.",
         f"{n_isd} ISDs -- {n_isd_elected} elected ({n_isd_elected - n_isd_gap} fully seeded, {n_isd_gap} still "
         f"need officeholder sourcing), {n_isd_appointed} appointed/excluded."),

        ("CCD",
         "Public community/junior college district enumeration for issue #14, with each district's legal name "
         "(per the Bond Review Board), county, website, and TASB Policy Online structure source.",
         f"{n_ccd} community college districts."),

        ("Special Districts (Non-MUD)",
         "Special-purpose district enumeration for issue #15 (all non-MUD types -- ESDs, hospital districts, "
         "groundwater/water districts, etc.), sourced from the Texas Comptroller's Special Purpose District "
         "Public Information Database (SPDPID). Excludes Municipal Utility Districts (tracked separately, "
         "issue #11) and Soil & Water Conservation Districts (issue #27, out of scope -- see that sheet).",
         f"{n_spd} districts, {n_spd_active} ACTIVE per their latest Comptroller filing."),

        ("MUD",
         "Municipal Utility District enumeration for issue #11, sourced from the same Comptroller SPDPID "
         "database as the Special Districts sheet but tracked on its own sheet per that issue's scope.",
         f"{n_mud} MUDs, {n_mud_active} ACTIVE per their latest Comptroller filing."),

        ("Judiciary",
         "Elected judiciary for issue #26: District Courts, Courts of Appeals, Supreme Court of Texas, and "
         "Court of Criminal Appeals, sourced from the Texas SOS 2024 General Election Winner Listing Report "
         "(2024-cycle winners only -- staggered terms mean an office absent here was not on the 2024 ballot, "
         "not vacant). County/Jurisdiction Served and Jurisdiction ID were backfilled from the judicial-district "
         "and appellate-district jurisdictions minted for #26 "
         "(data/us/tx/jurisdictions/{judicial-district,appellate-district,state}/); Party (2024) from the "
         "Official Canvass Report. County Court at Law, Probate Court, and Justice of the Peace rows are also "
         "included here even though their office/officeholder records live under the county layer "
         "(data/us/tx/.../county/), since they're court-adjacent. \"2ND MULTICOUNTY COURT AT LAW\" (Bee, Live "
         "Oak, McMullen Counties) has a jurisdiction confirmed via Legislature bill materials rather than a "
         "pinned Gov't Code section number -- see that row's Notes.",
         f"{n_jud} rows -- {jud_summary}."),

        ("Soil & Water Conservation",
         "OUT OF SCOPE for this project (issue #27, closed wontfix/Out of Scope): SWCD director elections are "
         "not open to the general public -- only landowners within the district vote, unlike every other entity "
         "type in this workbook. Sheet enumerates districts from the Texas State Soil and Water Conservation "
         "Board (tsswcb.texas.gov) directory and is kept for historical/reference purposes only; not being "
         "actively maintained, and no office/officeholder data will be built for it under this project.",
         f"{n_swcd} districts (as of research date; not maintained going forward)."),

        ("Appraisal Districts",
         "County Appraisal District tracking for issue #28: whether each of the 254 CADs has the 3 popularly-"
         "elected board seats that Tax Code Sec. 6.0301 adds for counties at or above the 2020 Census's "
         "75,000-population threshold (50 qualifying counties, collapsing to 49 rows since Potter and Randall "
         "share one joint CAD) -- alongside the standard 5 appointed + 1 ex officio (county Tax Assessor-"
         "Collector) board that every other CAD has.",
         f"{n_apd} CADs, {n_apd_elected} with the 3 elected seats."),
    ]


def main():
    write = "--write" in sys.argv[1:]

    wb = openpyxl.load_workbook(XLSX_PATH)
    rows = build_rows(wb)

    if "README" in wb.sheetnames:
        del wb["README"]
    ws = wb.create_sheet("README", 0)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 45

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "TX_Municipalities.xlsx -- Sheet Guide"
    ws["A1"].font = title_font
    ws["A2"] = ("Master reference workbook for CivicMirror's Texas elected-office audit (issue #10 and its "
                "follow-on issues, listed per sheet below). Regenerate this sheet with "
                "scripts/add_tx_municipalities_readme_sheet.py after updating any other sheet's row counts.")
    ws["A2"].alignment = wrap
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 30

    header_row = 4
    headers = ["Sheet", "What it tracks", "Coverage"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font

    for i, (sheet, desc, scope) in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=1, value=sheet).alignment = wrap
        ws.cell(row=i, column=2, value=desc).alignment = wrap
        ws.cell(row=i, column=3, value=scope).alignment = wrap
        ws.row_dimensions[i].height = 60

    print(f"README sheet built with {len(rows)} entries.")
    for sheet, _, scope in rows:
        print(f"  {sheet}: {scope}")

    if write:
        wb.save(XLSX_PATH)
        print(f"\nSaved {XLSX_PATH}")
    else:
        print("\n(dry run -- pass --write to save)")


if __name__ == "__main__":
    main()
