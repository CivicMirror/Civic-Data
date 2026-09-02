#!/usr/bin/env python3
"""
Fill in the "Judiciary" sheet of the TX_Municipalities.xlsx master workbook
with the jurisdiction data enumerated for issue #26 -- county/jurisdiction
served, a Jurisdiction ID cross-reference column, and party (now available
via the canvass join used in the officer import).

The sheet was originally populated straight from the 2024 SOS Winner
Listing Report, before any judicial-district/appellate-district
jurisdiction existed -- District Court and Court of Appeals rows have a
blank "County/Jurisdiction Served" column for that reason. This script
closes that gap using:
  - reference/TX Rolling Audit/OCA Judicial Directory/*.json (district and
    appellate county data, same sources as build_tx_judicial_jurisdictions.py)
  - the minted jurisdiction files under data/us/tx/jurisdictions/
  - the Official Canvass Report, for party

Updates the sheet in place; does not touch other sheets. Modifies the xlsx
directly -- run once, then commit the workbook.

Usage: python3 update_tx_municipalities_judiciary_sheet.py [--write]
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
import yaml

REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data" / "us" / "tx"
JD_SRC_DIR = REPO / "reference" / "TX Rolling Audit" / "OCA Judicial Directory"
SOS_SRC_DIR = REPO / "reference" / "TX Rolling Audit" / "2024 SOS Reports"
XLSX_PATH = REPO / "reference" / "TX Rolling Audit" / "TX_Municipalities.xlsx"

FIXUPS = {"dewitt": "de-witt", "lasalle": "la-salle"}


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def county_slug(raw):
    key = raw.strip().lower().replace(" ", "")
    return FIXUPS.get(key, slugify(raw))


def normalize_district_num(s):
    s = s.upper().strip()
    s = re.sub(r"(\d+)(ST|ND|RD|TH)\b", r"\1", s)
    return re.sub(r"\s+", " ", s)


ORDINAL_SLUG = {
    1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th", 6: "6th", 7: "7th",
    8: "8th", 9: "9th", 10: "10th", 11: "11th", 12: "12th", 13: "13th", 14: "14th",
}


def load_county_jurisdictions():
    out = {}
    for f in (DATA_DIR / "jurisdictions" / "county").glob("*-government.yaml"):
        slug = f.stem[: -len("-government")]
        doc = yaml.safe_load(f.read_text())
        out[slug] = doc["id"]
    return out


def load_judicial_district_map(known_counties):
    """normalized num -> (jurisdiction_id, county_list_str)"""
    raw = json.loads((JD_SRC_DIR / "district-courts-by-county-2026-05-01.json").read_text())
    by_slug = {f.stem: yaml.safe_load(f.read_text())["id"]
               for f in (DATA_DIR / "jurisdictions" / "judicial-district").glob("*.yaml")}
    out = {}
    for entry in raw:
        counties = entry["counties"]
        num = normalize_district_num(entry["num"])
        county_str = ", ".join(c.upper() for c in counties)
        if len(counties) == 1:
            jid = known_counties[county_slug(counties[0])]
        else:
            jid = by_slug["-".join(county_slug(c) for c in counties)]
        out[num] = (jid, county_str)
    return out


def load_appellate_district_map():
    raw = json.loads((JD_SRC_DIR / "appellate-districts-by-county-govcode-22.201.json").read_text())
    out = {}
    for num_str, counties in raw.items():
        num = int(num_str)
        if num == 15 or counties == "STATEWIDE":
            continue
        slug = ORDINAL_SLUG[num]
        doc = yaml.safe_load((DATA_DIR / "jurisdictions" / "appellate-district" / f"{slug}.yaml").read_text())
        out[num] = (doc["id"], ", ".join(c.upper() for c in counties))
    return out


def strip_incumbency(name):
    return re.sub(r"\s*\([A-Z]\)\s*$", "", name).strip()


def build_canvass_party_index():
    wb = openpyxl.load_workbook(SOS_SRC_DIR / "OfficialCanvassReport.xlsx")
    ws = wb.active
    idx = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        office, candidate, party = row[1], row[2], row[3]
        if not office or not candidate:
            continue
        idx[(office.strip(), strip_incumbency(candidate.strip()).upper())] = party.strip() if party else None
    return idx


STATE_JUDICIARY_ID = "ocd-jurisdiction/country:us/state:tx/judiciary"


def main():
    write = "--write" in sys.argv[1:]

    known_counties = load_county_jurisdictions()
    jd_map = load_judicial_district_map(known_counties)
    coa_map = load_appellate_district_map()
    party_idx = build_canvass_party_index()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Judiciary"]

    COL_COUNTY = 3   # C: County/Jurisdiction Served
    COL_PARTY = 6    # F: Party (2024)
    COL_NOTES = 7    # G: Notes
    COL_JID = 8      # H: Jurisdiction ID (new)

    ws.cell(row=2, column=COL_JID, value="Jurisdiction ID")

    stats = {"district_court": 0, "coa": 0, "coa_chief": 0, "statewide": 0,
             "party_filled": 0, "unresolved": 0, "already_had_county": 0}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        name_cell, type_cell = row[0], row[1]
        court_name, court_type = name_cell.value, type_cell.value
        if not court_name:
            continue
        court_name_norm = re.sub(r"\s*-\s*UNEXPIRED TERM\s*$", "", court_name).strip()

        jid, county_str = None, None

        if court_type == "District Court":
            m = re.match(r"^DISTRICT JUDGE,\s*(.+?)\s*JUDICIAL DISTRICT$", court_name_norm)
            if m:
                num = normalize_district_num(m.group(1))
                if num in jd_map:
                    jid, county_str = jd_map[num]
                    stats["district_court"] += 1
                else:
                    stats["unresolved"] += 1

        elif court_type in ("Court of Appeals", "Court of Appeals (Chief Justice)"):
            m = re.match(r"^(?:CHIEF )?JUSTICE,\s*(\d+)(?:ST|ND|RD|TH)\s*COURT OF APPEALS DISTRICT", court_name)
            if m:
                num = int(m.group(1))
                if num in coa_map:
                    jid, county_str = coa_map[num]
                    stats["coa_chief" if court_type.endswith("(Chief Justice)") else "coa"] += 1
                else:
                    stats["unresolved"] += 1

        elif court_type in ("Supreme Court of Texas", "Court of Criminal Appeals"):
            jid, county_str = STATE_JUDICIARY_ID, "Statewide"
            stats["statewide"] += 1

        else:
            # County Court at Law, Probate Court, Justice of the Peace already
            # carry a county in column C from the original SOS-derived pass;
            # leave that column alone, just backfill the jurisdiction id when
            # the county resolves to an existing seeded county jurisdiction.
            existing_county = row[COL_COUNTY - 1].value
            if existing_county:
                stats["already_had_county"] += 1
                slug = county_slug(existing_county.replace("COUNTY", "").strip())
                if slug in known_counties:
                    jid = known_counties[slug]

        if county_str and not row[COL_COUNTY - 1].value:
            ws.cell(row=row[0].row, column=COL_COUNTY, value=county_str)
        if jid:
            ws.cell(row=row[0].row, column=COL_JID, value=jid)

        if not row[COL_PARTY - 1].value:
            winner_cell = row[4].value  # E: 2024 Winner
            if winner_cell:
                party = party_idx.get((court_name, strip_incumbency(winner_cell).upper()))
                if party:
                    ws.cell(row=row[0].row, column=COL_PARTY, value=party)
                    stats["party_filled"] += 1

        if court_type == "Multicounty Court at Law" and not row[COL_NOTES - 1].value:
            ws.cell(row=row[0].row, column=COL_NOTES,
                    value="No jurisdiction minted yet -- multicounty court-at-law county composition not yet researched (issue #26).")

    print("==================== SUMMARY ====================")
    for k in sorted(stats):
        print(f"{k}: {stats[k]}")

    if write:
        wb.save(XLSX_PATH)
        print(f"\nSaved {XLSX_PATH}")
    else:
        print("\n(dry run -- pass --write to save)")


if __name__ == "__main__":
    main()
