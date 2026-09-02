#!/usr/bin/env python3
"""
Backfill the "ISD" sheet of TX_Municipalities.xlsx with the structure and
officeholder-coverage data merged into the repo for issue #13 (the
tx-audit-instructions branch's ISD work: 1,012 jurisdictions, 1,007
organizations, 1,809 posts, 5,600 people/memberships under
data/us/tx/{...}/school/).

Adds four columns, matched to each row by the sheet's existing "TEA CDN"
column (which already lines up with the tea-cdn external identifier on
each Organization record):
  - Jurisdiction ID
  - Board Size (seats)
  - Seat Structure (At-Large / Single-Member District / Hybrid / Zones,
    inferred from the post id/title pattern for that district)
  - Officeholders Seeded (e.g. "7/7" or "2/6 -- needs sourcing")
  - Status (Elected / Appointed (excluded) -- Notes column gets the reason
    for excluded districts, sourced from
    tx_isd_appointed_exclusions_2026-08-25.csv)

Districts with no matching organization (appointed exclusions) get
Status = "Appointed (excluded)" and a Notes entry; everything else is
"Elected".

Usage: python3 update_tx_municipalities_isd_sheet.py [--write]
"""
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml

REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data" / "us" / "tx"
XLSX_PATH = REPO / "reference" / "TX Rolling Audit" / "TX_Municipalities.xlsx"
EXCLUSIONS_CSV = REPO / "reference" / "TX Rolling Audit" / "tx_isd_appointed_exclusions_2026-08-25.csv"


def load_appointed_exclusions():
    out = {}
    with EXCLUSIONS_CSV.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["tea_cdn"]] = row["note"]
    return out


def load_orgs_by_cdn():
    out = {}
    for f in (DATA_DIR / "organizations" / "school").glob("*.yaml"):
        doc = yaml.safe_load(f.read_text())
        for ident in doc.get("identifiers", []):
            if ident["scheme"] == "tea-cdn":
                out[ident["identifier"]] = doc
    return out


def load_posts_by_org():
    out = defaultdict(list)
    for f in (DATA_DIR / "posts" / "school").glob("*.yaml"):
        doc = yaml.safe_load(f.read_text())
        out[doc["organization_id"]].append(doc)
    return out


def load_membership_counts_by_post():
    out = defaultdict(int)
    for f in (DATA_DIR / "memberships" / "school").glob("*.yaml"):
        doc = yaml.safe_load(f.read_text())
        out[doc["post_id"]] += 1
    return out


def classify_structure(posts):
    ids = [p["id"] for p in posts]
    has_at_large = any(re.search(r"trustee(-at-large)?$", pid) and "district" not in pid and "zone" not in pid and "precinct" not in pid for pid in ids)
    has_district = any("district-" in pid for pid in ids)
    has_zone = any("zone" in pid for pid in ids)
    has_precinct = any("precinct" in pid for pid in ids)
    if has_zone:
        return "Hybrid (zones)" if has_at_large else "Zones"
    if has_precinct:
        return "Elected precincts"
    if has_district and has_at_large:
        return "Hybrid (at-large + SMD)"
    if has_district:
        return "Single-Member District"
    return "At-Large"


def main():
    write = "--write" in sys.argv[1:]

    exclusions = load_appointed_exclusions()
    orgs_by_cdn = load_orgs_by_cdn()
    posts_by_org = load_posts_by_org()
    membership_counts = load_membership_counts_by_post()

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["ISD"]

    COL_CDN = 3       # C: TEA CDN
    COL_NOTES = 6     # F: Notes
    COL_JID = 7       # G: Jurisdiction ID (new)
    COL_SEATS = 8     # H: Board Size (new)
    COL_STRUCTURE = 9  # I: Seat Structure (new)
    COL_COVERAGE = 10  # J: Officeholders Seeded (new)
    COL_STATUS = 11    # K: Status (new)

    headers = {
        COL_JID: "Jurisdiction ID",
        COL_SEATS: "Board Size",
        COL_STRUCTURE: "Seat Structure",
        COL_COVERAGE: "Officeholders Seeded",
        COL_STATUS: "Status",
    }
    for col, label in headers.items():
        ws.cell(row=1, column=col, value=label)

    stats = {"elected": 0, "appointed_excluded": 0, "no_match": 0, "fully_seeded": 0, "needs_sourcing": 0}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cdn_cell = row[COL_CDN - 1]
        cdn = str(cdn_cell.value).strip().zfill(6) if cdn_cell.value else None
        if not cdn:
            continue

        if cdn in exclusions:
            ws.cell(row=row[0].row, column=COL_STATUS, value="Appointed (excluded)")
            if not row[COL_NOTES - 1].value:
                ws.cell(row=row[0].row, column=COL_NOTES, value=exclusions[cdn])
            stats["appointed_excluded"] += 1
            continue

        org = orgs_by_cdn.get(cdn)
        if not org:
            stats["no_match"] += 1
            continue

        jurisdiction_id = org["jurisdiction_id"]
        posts = posts_by_org.get(org["id"], [])
        total_seats = sum(p.get("seats", 1) for p in posts)
        total_seated = sum(min(membership_counts.get(p["id"], 0), p.get("seats", 1)) for p in posts)
        structure = classify_structure(posts)

        ws.cell(row=row[0].row, column=COL_JID, value=jurisdiction_id)
        ws.cell(row=row[0].row, column=COL_SEATS, value=total_seats)
        ws.cell(row=row[0].row, column=COL_STRUCTURE, value=structure)
        coverage = f"{total_seated}/{total_seats}"
        if total_seated < total_seats:
            coverage += " -- needs sourcing"
            stats["needs_sourcing"] += 1
        else:
            stats["fully_seeded"] += 1
        ws.cell(row=row[0].row, column=COL_COVERAGE, value=coverage)
        ws.cell(row=row[0].row, column=COL_STATUS, value="Elected")
        stats["elected"] += 1

    print("==================== SUMMARY ====================")
    for k in sorted(stats):
        print(f"{k}: {stats[k]}")

    if write:
        wb.save(XLSX_PATH)
        print(f"\nSaved {XLSX_PATH}")
    else:
        print("\n(dry run -- pass --write to save)")


if __name__ == "__main__":
    main()
