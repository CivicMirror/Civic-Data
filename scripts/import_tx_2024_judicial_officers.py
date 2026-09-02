#!/usr/bin/env python3
"""
Import TX judicial officers (District Judges, District Attorneys, Criminal
District Judges, Courts of Appeals Justices, Supreme Court Justices, Court
of Criminal Appeals Judges) from the 2024 SOS Winner Listing Report into
data/us/tx/{organizations,posts,memberships,people}/judicial/, for issue #26.

Companion to scripts/import_tx_2024_county_officers.py (#29, county layer)
and scripts/build_tx_judicial_jurisdictions.py (#26, jurisdiction layer) --
run after both.

Jurisdiction resolution:
  - Single-county judicial districts (400 of 497) and the 3 counties with
    their own numbered "Criminal District Court"s (Dallas, Tarrant, El
    Paso) reuse the existing county jurisdiction -- no new jurisdiction was
    minted for these, per the #29 precedent for County Court at Law judges.
  - Multi-county judicial districts, the Kleberg/Kenedy DA district, and
    the 14 Courts of Appeals districts use the jurisdictions minted by
    build_tx_judicial_jurisdictions.py.
  - Supreme Court, Court of Criminal Appeals: state:tx/judiciary
    (data/us/tx/jurisdictions/state/tx-judiciary.yaml).

Unlike the county layer (#29), party IS available here: the Official
Canvass Report does cover every one of these race types (District Judge,
DA, Courts of Appeals, Supreme Court, CCA -- 377 races total per the #26
research comments), so this script joins on (office, candidate name) to
recover party per candidacy.

Currency caveat, same as #29: these are 2024-cycle winners. District Court/
DA seats not on the 2024 ballot (staggered terms; a few "UNEXPIRED TERM"
seats besides) are simply absent here, not vacant.

Idempotent by id and filename. Dry run by default; pass --write to persist.
"""
import json
import re
import sys
import unicodedata
import uuid
from pathlib import Path

import openpyxl
import yaml

REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data" / "us" / "tx"
SRC_DIR = REPO / "reference" / "TX Rolling Audit" / "2024 SOS Reports"
JD_SRC_DIR = REPO / "reference" / "TX Rolling Audit" / "OCA Judicial Directory"
WINNERS_XLSX = SRC_DIR / "WinnerListingReport.xlsx"
CANVASS_XLSX = SRC_DIR / "OfficialCanvassReport.xlsx"

RESULTS_URL = "https://results.texas-election.com/reports"
RETRIEVED = "2026-09-01"
ELECTION_ID = "us-tx/2024-11-05/general"

NS = uuid.UUID("3b6f0a9e-9a3b-5f2e-8b8f-6c1b6e9d2a41")  # shared with import_tx_2024_county_officers.py

FIXUPS = {"dewitt": "de-witt", "lasalle": "la-salle"}


def slugify(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def county_slug(raw):
    key = raw.strip().lower().replace(" ", "")
    return FIXUPS.get(key, slugify(raw))


def oid(kind, key):
    return f"ocd-{kind}/{uuid.uuid5(NS, kind + '|' + key)}"


ORDINAL_SLUG = {
    1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th", 6: "6th", 7: "7th",
    8: "8th", 9: "9th", 10: "10th", 11: "11th", 12: "12th", 13: "13th", 14: "14th",
}
ORDINAL_NAME = {
    1: "First", 2: "Second", 3: "Third", 4: "Fourth", 5: "Fifth", 6: "Sixth", 7: "Seventh",
    8: "Eighth", 9: "Ninth", 10: "Tenth", 11: "Eleventh", 12: "Twelfth", 13: "Thirteenth", 14: "Fourteenth",
}


def ordinal(n):
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def normalize_district_num(s):
    """Collapse a district-number string to a comparison key, independent of
    whether it carries an ordinal suffix: OCA's raw '100' and the SOS
    winner file's '100TH' (or '2nd 25' vs '2ND 25TH') must resolve to the
    same jurisdiction lookup key."""
    s = s.upper().strip()
    s = re.sub(r"(\d+)(ST|ND|RD|TH)\b", r"\1", s)
    return re.sub(r"\s+", " ", s)


def lower_ordinal_suffix(s):
    """SOS-style '100TH' / '2ND 25TH' -> '100th' / '2nd 25th' for display."""
    return re.sub(r"(\d+)(ST|ND|RD|TH)\b", lambda m: m.group(1) + m.group(2).lower(), s)


def titlecase_name(raw):
    return " ".join(w.capitalize() for w in raw.split())


# --- Jurisdiction resolution ------------------------------------------------

def load_county_jurisdictions():
    out = {}
    for f in (DATA_DIR / "jurisdictions" / "county").glob("*-government.yaml"):
        slug = f.stem[: -len("-government")]
        doc = yaml.safe_load(f.read_text())
        out[slug] = doc["id"]
    return out


def load_judicial_district_number_map(known_counties):
    """num (e.g. '18', '1A', '2nd 25') -> (jurisdiction_id, org_slug)."""
    raw = json.loads((JD_SRC_DIR / "district-courts-by-county-2026-05-01.json").read_text())
    jd_dir = {}
    for f in (DATA_DIR / "jurisdictions" / "judicial-district").glob("*.yaml"):
        doc = yaml.safe_load(f.read_text())
        jd_dir[doc["id"]] = doc
    by_slug = {f.stem: yaml.safe_load(f.read_text())["id"]
               for f in (DATA_DIR / "jurisdictions" / "judicial-district").glob("*.yaml")}

    out = {}
    for entry in raw:
        counties = entry["counties"]
        num = normalize_district_num(entry["num"])
        if len(counties) == 1:
            slug = county_slug(counties[0])
            out[num] = (known_counties[slug], f"{slug}-tx-county")
        else:
            combo_slug = "-".join(county_slug(c) for c in counties)
            out[num] = (by_slug[combo_slug], f"{combo_slug}-tx-judicial-district")
    return out


def load_appellate_district_map():
    out = {}
    for n, slug in ORDINAL_SLUG.items():
        f = DATA_DIR / "jurisdictions" / "appellate-district" / f"{slug}.yaml"
        doc = yaml.safe_load(f.read_text())
        out[n] = (doc["id"], f"{slug}-tx-appellate-district")
    return out


STATE_JUDICIARY_ID = "ocd-jurisdiction/country:us/state:tx/judiciary"
STATE_ORG_SLUG = "tx-state"


# --- Winner / canvass loading ------------------------------------------------

def load_rows(path, skip_title_row):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if skip_title_row:
        rows = rows[1:]
    return rows[1:]


def strip_incumbency(name):
    return re.sub(r"\s*\([A-Z]\)\s*$", "", name).strip()


def build_canvass_party_index():
    rows = load_rows(CANVASS_XLSX, skip_title_row=False)
    idx = {}
    for row in rows:
        office, candidate, party = row[1], row[2], row[3]
        if not office or not candidate:
            continue
        key = (office.strip(), strip_incumbency(candidate.strip()).upper())
        idx[key] = party.strip() if party else None
    return idx


# --- Recorder ------------------------------------------------------------

class Recorder:
    def __init__(self, write):
        self.write = write
        self.stats = {}

    def emit(self, kind, doc, filename, header):
        singular = kind.rstrip("s")
        path = DATA_DIR / kind / "judicial" / filename
        if path.exists():
            existing = yaml.safe_load(path.read_text()) or {}
            if existing.get("id") == doc["id"]:
                self.stats[f"{singular}_existing"] = self.stats.get(f"{singular}_existing", 0) + 1
            else:
                self.stats[f"{singular}_skipped_filename_conflict"] = (
                    self.stats.get(f"{singular}_skipped_filename_conflict", 0) + 1
                )
            return False
        self.stats[f"{singular}_new"] = self.stats.get(f"{singular}_new", 0) + 1
        if self.write:
            path.parent.mkdir(parents=True, exist_ok=True)
            text = header + yaml.safe_dump(
                doc, sort_keys=False, allow_unicode=True, default_flow_style=False, width=1000
            )
            path.write_text(text)
        return True


HEADER = (
    "# Imported from the Texas Secretary of State's 2024 General Election\n"
    "# Winner Listing Report (results.texas-election.com/reports), joined\n"
    "# with the Official Canvass Report for party, by scripts/\n"
    "# import_tx_2024_judicial_officers.py. See issue #26. Offices not on\n"
    "# the 2024 ballot (staggered terms) are simply not represented here,\n"
    "# not vacant -- see the script's module docstring.\n"
)


def add_record(rec, orgs, posts, people, memberships, *, jurisdiction_id, org_slug, org_key,
               org_name, post_key, post_title, seats, candidate_raw, party, county_hint_slug=None):
    full_org_key = (org_slug, org_key)
    post_id = f"{org_slug}/{post_key}"

    if full_org_key not in orgs:
        org_id = oid("organization", f"tx-judicial|{org_slug}|{org_key}")
        orgs[full_org_key] = {
            "id": org_id,
            "name": org_name,
            "jurisdiction_id": jurisdiction_id,
            "identifiers": [{"scheme": "civicmirror-organization", "identifier": f"{org_slug}/{org_key}"}],
            "status": "active",
            "sources": [{"url": RESULTS_URL,
                         "note": "Texas SOS 2024 General Election Winner Listing Report",
                         "retrieved": RETRIEVED}],
        }

    org_id = orgs[full_org_key]["id"]

    # post creation is keyed independently of org creation: several distinct
    # posts (e.g. Places 2/4/6 on the Supreme Court) share one organization.
    if post_id not in posts:
        posts[post_id] = {
            "id": post_id,
            "organization_id": org_id,
            "title": post_title,
            "seats": seats,
            "identifiers": [{"scheme": "civicmirror-office", "identifier": post_id}],
            "sources": [{"url": RESULTS_URL,
                         "note": "Texas SOS 2024 General Election Winner Listing Report",
                         "retrieved": RETRIEVED}],
        }
    candidate_name = re.sub(r"\s+", " ", candidate_raw).strip()
    person_id = oid("person", f"tx-judicial-2024|{post_id}|{candidate_name}")

    person = {
        "id": person_id,
        "name": titlecase_name(candidate_name),
        "candidacies": [{
            "contest_id": f"tx-2024-general/{post_id}",
            "election_id": ELECTION_ID,
            "jurisdiction_id": jurisdiction_id,
            "office_id": post_id,
            "party": party,
            "ballot_name": candidate_name,
            "status": "active",
            "sources": [{"url": RESULTS_URL,
                         "note": "Texas SOS 2024 General Election Winner Listing Report; party from the Official Canvass Report.",
                         "retrieved": RETRIEVED}],
        }],
        "verification": {"status": "machine-extracted", "pipeline": "import_tx_2024_judicial_officers"},
        "sources": [{"url": RESULTS_URL,
                     "note": "Texas SOS 2024 General Election Winner Listing Report",
                     "retrieved": RETRIEVED}],
    }
    people.append(person)

    surname = candidate_name.split()[-1]
    mem_key = f"{post_id.replace('/', '-')}-{slugify(surname)}"
    note = ("2024 General Election winner; office serves a staggered term, so this "
            "reflects the seat as of the 2024 cycle, not a verified 2026 incumbency check.")
    membership = {
        "id": mem_key,
        "person_id": person_id,
        "organization_id": org_id,
        "post_id": post_id,
        "role": post_title,
        "how_seated": "elected",
        "end": "2029",
        "sources": [{"url": RESULTS_URL, "note": note, "retrieved": RETRIEVED}],
    }
    memberships.append((mem_key, membership))


def main():
    write = "--write" in sys.argv[1:]

    known_counties = load_county_jurisdictions()
    jd_by_num = load_judicial_district_number_map(known_counties)
    coa_by_num = load_appellate_district_map()

    kleberg_kenedy_jurisdiction = None
    kk_path = DATA_DIR / "jurisdictions" / "judicial-district" / "kenedy-kleberg.yaml"
    kleberg_kenedy_jurisdiction = yaml.safe_load(kk_path.read_text())["id"]

    party_idx = build_canvass_party_index()

    winner_rows = load_rows(WINNERS_XLSX, skip_title_row=True)

    orgs, posts, people, memberships = {}, {}, [], []
    unmatched = {}

    def party_for(office_raw, candidate_raw):
        return party_idx.get((office_raw, strip_incumbency(candidate_raw).upper()))

    for row in winner_rows:
        office_raw, candidate_raw = row[1], row[2]
        if not office_raw or not candidate_raw:
            continue
        office_raw = office_raw.strip()
        office = re.sub(r"\s*-\s*UNEXPIRED TERM\s*$", "", office_raw).strip()
        party = party_for(office_raw, candidate_raw)

        m = re.match(r"^DISTRICT JUDGE,\s*(.+?)\s*JUDICIAL DISTRICT$", office)
        if m:
            raw_num = m.group(1).strip()
            num = normalize_district_num(raw_num)
            if num not in jd_by_num:
                unmatched[office_raw] = unmatched.get(office_raw, 0) + 1
                continue
            jid, org_slug = jd_by_num[num]
            label = lower_ordinal_suffix(raw_num)
            slug = slugify(num)
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=jid, org_slug=org_slug, org_key=f"district-judge-{slug}",
                       org_name=f"{label} District Court", post_key=f"district-judge-{slug}",
                       post_title=f"District Judge, {label} Judicial District", seats=1,
                       candidate_raw=candidate_raw, party=party)
            continue

        m = re.match(r"^DISTRICT ATTORNEY,\s*(.+?)\s*JUDICIAL DISTRICT$", office)
        if m:
            raw_num = m.group(1).strip()
            num = normalize_district_num(raw_num)
            if num not in jd_by_num:
                unmatched[office_raw] = unmatched.get(office_raw, 0) + 1
                continue
            jid, org_slug = jd_by_num[num]
            label = lower_ordinal_suffix(raw_num)
            slug = slugify(num)
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=jid, org_slug=org_slug, org_key=f"district-attorney-{slug}",
                       org_name=f"District Attorney, {label} Judicial District",
                       post_key=f"district-attorney-{slug}",
                       post_title=f"District Attorney, {label} Judicial District", seats=1,
                       candidate_raw=candidate_raw, party=party)
            continue

        if office == "DISTRICT ATTORNEY FOR KLEBERG AND KENEDY COUNTIES":
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=kleberg_kenedy_jurisdiction, org_slug="kenedy-kleberg-tx-judicial-district",
                       org_key="district-attorney", org_name="District Attorney for Kleberg and Kenedy Counties",
                       post_key="district-attorney", post_title="District Attorney for Kleberg and Kenedy Counties",
                       seats=1, candidate_raw=candidate_raw, party=party)
            continue

        if office == "HARRIS COUNTY DISTRICT ATTORNEY":
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=known_counties["harris"], org_slug="harris-tx-county",
                       org_key="district-attorney", org_name="Harris County District Attorney",
                       post_key="district-attorney", post_title="Harris County District Attorney",
                       seats=1, candidate_raw=candidate_raw, party=party)
            continue

        m = re.match(r"^CRIMINAL DISTRICT JUDGE,\s*(.+?)\s*COUNTY NUMBER\s*(\d+)$", office)
        m2 = re.match(r"^CRIMINAL DISTRICT JUDGE #(\d+)\s*(.+?)\s*COUNTY$", office)
        if m or m2:
            if m:
                county_name, n = m.group(1), m.group(2)
            else:
                n, county_name = m2.group(1), m2.group(2)
            slug = county_slug(county_name)
            if slug not in known_counties:
                unmatched[office_raw] = unmatched.get(office_raw, 0) + 1
                continue
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=known_counties[slug], org_slug=f"{slug}-tx-county",
                       org_key=f"criminal-district-judge-number-{n}",
                       org_name=f"Criminal District Court Number {n} ({county_name.strip().title()} County)",
                       post_key=f"criminal-district-judge-number-{n}",
                       post_title=f"Criminal District Judge, Number {n}", seats=1,
                       candidate_raw=candidate_raw, party=party)
            continue

        m = re.match(r"^(CHIEF )?JUSTICE,\s*(\d+)(?:ST|ND|RD|TH)\s*COURT OF APPEALS DISTRICT(?:,\s*PLACE\s*(\d+))?$", office)
        if m:
            is_chief, num_str, place = m.group(1), m.group(2), m.group(3)
            num = int(num_str)
            if num not in coa_by_num:
                unmatched[office_raw] = unmatched.get(office_raw, 0) + 1
                continue
            jid, org_slug = coa_by_num[num]
            coa_name = f"{ORDINAL_NAME[num]} Court of Appeals"
            if is_chief:
                add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                           jurisdiction_id=jid, org_slug=org_slug, org_key="chief-justice",
                           org_name=coa_name, post_key="chief-justice",
                           post_title=f"Chief Justice, {coa_name}", seats=1,
                           candidate_raw=candidate_raw, party=party)
            else:
                add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                           jurisdiction_id=jid, org_slug=org_slug, org_key=f"justice-place-{place}",
                           org_name=coa_name, post_key=f"justice-place-{place}",
                           post_title=f"Justice, {coa_name}, Place {place}", seats=1,
                           candidate_raw=candidate_raw, party=party)
            continue

        m = re.match(r"^JUSTICE,\s*SUPREME COURT,\s*PLACE\s*(\d+)$", office)
        if m:
            place = m.group(1)
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=STATE_JUDICIARY_ID, org_slug=STATE_ORG_SLUG,
                       org_key="scotx", org_name="Supreme Court of Texas",
                       post_key=f"scotx-justice-place-{place}",
                       post_title=f"Justice, Supreme Court of Texas, Place {place}", seats=1,
                       candidate_raw=candidate_raw, party=party)
            continue

        m = re.match(r"^JUDGE,\s*COURT OF CRIMINAL APPEALS,\s*PLACE\s*(\d+)$", office)
        if m:
            place = m.group(1)
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=STATE_JUDICIARY_ID, org_slug=STATE_ORG_SLUG,
                       org_key="cca", org_name="Court of Criminal Appeals",
                       post_key=f"cca-judge-place-{place}",
                       post_title=f"Judge, Court of Criminal Appeals, Place {place}", seats=1,
                       candidate_raw=candidate_raw, party=party)
            continue

        if office == "PRESIDING JUDGE, COURT OF CRIMINAL APPEALS":
            add_record(rec=None, orgs=orgs, posts=posts, people=people, memberships=memberships,
                       jurisdiction_id=STATE_JUDICIARY_ID, org_slug=STATE_ORG_SLUG,
                       org_key="cca", org_name="Court of Criminal Appeals",
                       post_key="cca-presiding-judge",
                       post_title="Presiding Judge, Court of Criminal Appeals", seats=1,
                       candidate_raw=candidate_raw, party=party)
            continue

    rec = Recorder(write)
    for key, org in orgs.items():
        rec.emit("organizations", org, f"{org['id'].rsplit('/', 1)[-1]}.yaml", HEADER)
    for key, post in posts.items():
        rec.emit("posts", post, f"{post['id'].replace('/', '-')}.yaml", HEADER)

    taken_people_slugs = ({p.stem for p in (DATA_DIR / "people" / "judicial").glob("*.yaml")}
                           if (DATA_DIR / "people" / "judicial").exists() else set())
    for person in people:
        base_slug = slugify(person["name"])
        suffix = person["id"].rsplit("/", 1)[-1][:8]
        slug_name = base_slug if base_slug not in taken_people_slugs else f"{base_slug}-{suffix}"
        taken_people_slugs.add(slug_name)
        rec.emit("people", person, f"{slug_name}.yaml", HEADER)

    for mem_key, membership in memberships:
        rec.emit("memberships", membership, f"{mem_key}.yaml", HEADER)

    print(f"Parsed {len(winner_rows)} winner rows")
    print(f"Organizations: {len(orgs)}  Posts: {len(posts)}  People: {len(people)}  Memberships: {len(memberships)}")
    print()
    print("==================== SUMMARY ====================")
    for k in sorted(rec.stats):
        print(f"{k}: {rec.stats[k]}")
    if unmatched:
        print(f"\nUnmatched judicial office strings ({sum(unmatched.values())} rows, {len(unmatched)} distinct):")
        for k in sorted(unmatched):
            print(f"  {k!r}  x{unmatched[k]}")
    if not write:
        print("\n(dry run -- pass --write to create files)")


if __name__ == "__main__":
    main()
